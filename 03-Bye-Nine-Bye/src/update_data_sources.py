#!/usr/bin/python

import pandas as pd
import json
from urllib.request import Request, urlopen
from openpyxl import load_workbook

pd.io.formats.excel.ExcelFormatter.header_style = None


def get_from_api(url: str, source: str) -> pd.DataFrame:
    """DOWNLOAD DATA FROM URL"""
    
    if source == "json":
        resp = urlopen(url)
        data = json.loads(resp.read())
        df = pd.DataFrame.from_dict(data["dataSet"])
    elif source == "csv":
        resp = Request(url, headers={
            "User-Agent": (
                "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:77.0) "
                "Gecko/20100101 Firefox/77.0"
            )
        })
        data = urlopen(resp)
        df = pd.read_csv(data)
    
    return df


def update_excel_workbook(sheet_data: dict) -> None:
    """UPDATE EXCEL WORKSHEET"""
    
    excel_file = "./data/hk_population_movement.xlsx"
    
    book = load_workbook(excel_file)
    writer = pd.ExcelWriter(excel_file, engine="openpyxl")
    writer.book = book
    
    for sheet, data in sheet_data.items():
        idx = book.sheetnames.index(sheet)
        
        book.remove(book.worksheets[idx])
        book.create_sheet(sheet, idx)
        
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        data.to_excel(writer, sheet, index=0, startrow=0, startcol=0)
    
    writer.save()
    
    return None


def get_hk_population() -> pd.DataFrame:
    """DOWNLOAD AND WRANGLE POPULATION DATA"""
    
    # Table 1A: Population by Sex and Age Group
    url_population = (
        "https://www.censtatd.gov.hk/api/get.php?id=1A&lang=en&param=N4Igx"
        "gbiBcoMJwJqJqAjDEAGHu+4HYB9AT33IpABoQiAXIzLa2gB0ejQF8aAJVEBmggAzC"
        "3ods49jG40AygFMAHgKGjxDTBhpEZ0LDxABBAOaK1mItc2SdbDoaMBnKLAcwATDQA"
        "KAeR8YAG0QACUAQwB3IgBpIjQAE3ZWRQAnIgA7Fh9UgHt2RPYAUkyQAF0jABsYOlS"
        "AV0UuIA"
    )
    df_population = get_from_api(url_population, "json")
    
    df_population = df_population.query(
        "Sex == ''"
        "and Age == ''"
        "and svDesc == 'Number (\\'000)'"
    )
    
    col = ["H", "period", "figure"]
    df_population = df_population[col].assign(
        sv="Population"
    )
    
    # Table 2: Population Growth by Component
    url_component = (
        "https://www.censtatd.gov.hk/api/get.php?id=2&lang=en&param=N4Igxg"
        "biBcoMJwJqJqAjDEAGHu+4HYB9AT33IpABoQiAXIzLa2gB0eiwF8aAJVEBmggAzC3"
        "ods49jG406JVgFMiASwB2AMwA2AewDuAoaPENMGGkRnQ0XHiADOUWGw4AmGgCFVAJ"
        "zoALBysfXRgAbRAAJQBDfSIAaSI0ABN2ZR8idRAAXRoAESVogKDWEPCo2ISk1KslD"
        "KzckAA5AElymLjElLS6zJyaJoBZdsqumvS+xoAFAHERzuqe+v6QWZ95xaIAUkn7bR"
        "g6HwBXJS4gA"
    )
    df_component = get_from_api(url_component, "json")
    
    df_component = df_component.query(
        "sv == 'Births_pro'"
        "or sv == 'Deaths_pro'"
        "or (sv == 'NM' and type_inflow == 'one-way')"
        "or (sv == 'NM' and type_inflow == 'others')"
    ).copy()
    df_component["sv"] = df_component["type_inflowDesc"].fillna(
        df_component["sv"].str.replace("_pro", "")
    )
    
    col = ["H", "period", "figure", "sv"]
    df_component = df_component[col]
    
    # concatenate data frames
    df = pd.concat([df_population, df_component])
    
    df = df.pivot(index=["period", "H"], columns=["sv"], values="figure")
    
    df[["Deaths", "Population"]] *= -1
    df["Population PP"] = df.sum(axis=1)
    df[["Population", "Population PP"]] *= -1
    
    df = df.reset_index() \
        .melt(id_vars=["period", "H"], value_vars=df.columns) \
        .sort_values(by=["period", "H", "sv"], ignore_index=True) \
        .rename(columns={
            "period": "Year",
            "H": "Quarter",
            "sv": "Event",
            "value": "Population",
        }).copy()
    
    df["Year"] = df["Year"].astype(int)
    df["Quarter"] = df["Quarter"].astype(int) * 2
    df["Population"] = df["Population"].astype(int) * 1_000
    
    return df


def get_hk_traffic() -> pd.DataFrame:
    """DOWNLOAD AND WRANGLE PASSENGER TRAFFIC DATA"""
    
    # Statistics on Passenger Traffic
    url_traffic = (
        "https://www.immd.gov.hk/opendata/eng/transport/"
        "immigration_clearance/statistics_on_daily_passenger_traffic.csv"
    )
    df = get_from_api(url_traffic, "csv")
    
    df["Date"] = pd.to_datetime(df["Date"], format="%d-%m-%Y")
    df["Year"] = df["Date"].dt.year
    df["Quarter"] = df["Date"].dt.quarter
    df["Month"] = df["Date"].dt.month_name()
    
    clus = ["Year", "Quarter", "Month", "Arrival / Departure"]
    col = ["Hong Kong Residents", "Mainland Visitors", "Other Visitors"]
    df = df.groupby(clus)[col].sum().reset_index()
    
    df = df.melt(id_vars=clus, value_vars=col) \
        .sort_values(by=clus, ignore_index=True) \
        .rename(columns={
            "Arrival / Departure": "Movement",
            "variable": "Passenger",
            "value": "Traffic",
        })
    
    df.loc[df["Movement"] == "Departure", "Traffic"] *= -1
    
    return df


def update_data() -> None:
    """UPDATE AGGREGATED DATA INTO EXCEL FILE"""
    
    hk_population = get_hk_population()
    hk_traffic = get_hk_traffic()
    
    update_excel_workbook({
        "HK_POPULATION": hk_population,
        "HK_TRAFFIC": hk_traffic,
    })
    
    print("Finished update!")
    
    return None


if __name__ == "__main__":
    update_data()
